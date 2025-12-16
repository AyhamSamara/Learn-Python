import openpyxl
from openpyxl.styles import Font, PatternFill, Color

# ==========================================
# PART 1: CREATING A NEW EXCEL FILE
# ==========================================
print("--- 1. Creating a New Workbook ---")

# 1. Create a blank Workbook
wb = openpyxl.Workbook()

# 2. Grab the active sheet (the default first tab)
ws = wb.active
ws.title = "Student Grades"  # Rename the tab

# 3. Add Headers
headers = ["Student Name", "Math Score", "Science Score"]
ws.append(headers)  # .append() adds a list as a new row at the bottom

# 4. Add some Dummy Data
data = [
    ["Ayham", 85, 90],
    ["Tareq", 45, 60],
    ["Omar", 70, 55],
    ["Yousef", 95, 99],
    ["Zaid", 40, 35]
]

for row in data:
    ws.append(row)

# # 5. Save the file
filename = ".\\sample\\grades_report.xlsx"
wb.save(filename)
print(f"Created '{filename}' with dummy data.\n")

# ==========================================
# PART 2: READING & PROCESSING DATA
# ==========================================
print("--- 2. Processing Data (Calculating Totals) ---")

# 1. Load the workbook we just created
# data_only=True is important! It reads the *value* (e.g., 10), not the formula (e.g., =A1+B1)
wb = openpyxl.load_workbook(filename)
ws = wb["Student Grades"]  # Select the sheet by name

# 2. Add a new header for "Total" and "Status"
# We know headers are in Row 1.
# Cell(row, col) is easier for loops. Rows/Cols start at 1 (not 0).
ws.cell(row=1, column=4, value="Total Score")
ws.cell(row=1, column=5, value="Status")

# 3. Iterate through rows (Skipping the header row!)
# ws.max_row gives us the number of the last row with data.
for row_num in range(2, ws.max_row + 1):
    # Read values from Math (Col 2) and Science (Col 3)
    math_score = ws.cell(row=row_num, column=2).value
    science_score = ws.cell(row=row_num, column=3).value

    # Calculate Total
    total = math_score + science_score

    # Write Total to Column 4
    ws.cell(row=row_num, column=4, value=total)

    # Logic: Determine Pass/Fail
    # Let's say pass mark is 100 total
    status_cell = ws.cell(row=row_num, column=5)

    if total >= 100:
        status_cell.value = "PASS"
        # STYLING: Make text Green and Bold
        status_cell.font = Font(color="008000", bold=True)
    else:
        status_cell.value = "FAIL"
        # STYLING: Make text Red and Bold
        status_cell.font = Font(color="FF0000", bold=True)

    print(f"Processed row {row_num}: {ws.cell(row=row_num, column=1).value} -> {total}")
wb.save(filename)
# # ==========================================
# # PART 3: SAVING THE RESULT
# # ==========================================
new_filename = "grades_report_processed.xlsx"
wb.save(new_filename)
print(f"\n--- Done! Saved processed file as '{new_filename}' ---")